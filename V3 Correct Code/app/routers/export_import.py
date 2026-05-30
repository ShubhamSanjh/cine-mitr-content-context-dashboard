"""
Excel Export, Import, and Template endpoints.
Supports: Media, Links, Status Tracking data.
Uses openpyxl for .xlsx generation and parsing.
Includes: Import History tracking, Export errors, Scheduling.
"""

import io
import csv
from datetime import datetime, date as date_type, timedelta
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models.media import MediaContent, MediaLink, MediaStatus
from app.models.import_history import ImportHistory, ImportRecord, ImportSchedule

router = APIRouter(prefix="/excel")


def _parse_date(val):
    """Convert various date formats from Excel to Python date object."""
    if val is None:
        return None
    if isinstance(val, date_type):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s in ("None", ""):
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

# --- Styling helpers ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def workbook_to_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== EXPORT ====================

@router.get("/export/media", summary="Export all media to Excel")
def export_media_excel(
    media_category: str | None = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(MediaContent)
    if media_category:
        query = query.filter(MediaContent.media_category == media_category)
    items = query.order_by(MediaContent.updated_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Media"
    headers = ["ID", "Category", "Name", "Release Date", "Genre", "Director", "Cast", "Rating", "Review", "Tags", "Is Available", "Available On", "Created At"]
    style_header(ws, headers)

    for row, m in enumerate(items, 2):
        ws.cell(row=row, column=1, value=m.id)
        ws.cell(row=row, column=2, value=m.media_category)
        ws.cell(row=row, column=3, value=m.media_name)
        ws.cell(row=row, column=4, value=str(m.release_date) if m.release_date else "")
        ws.cell(row=row, column=5, value=m.genre or "")
        ws.cell(row=row, column=6, value=m.director or "")
        ws.cell(row=row, column=7, value=m.cast_members or "")
        ws.cell(row=row, column=8, value=m.rating)
        ws.cell(row=row, column=9, value=m.review or "")
        ws.cell(row=row, column=10, value=m.tags or "")
        ws.cell(row=row, column=11, value=m.is_available or "false")
        ws.cell(row=row, column=12, value=m.available_on or "")
        ws.cell(row=row, column=13, value=str(m.created_at) if m.created_at else "")

    auto_width(ws)
    return workbook_to_response(wb, f"media_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


@router.get("/export/links", summary="Export all links to Excel")
def export_links_excel(
    link_category: str | None = Query(None),
    link_status: str | None = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(MediaLink)
    if link_category:
        query = query.filter(MediaLink.link_category == link_category)
    if link_status:
        query = query.filter(MediaLink.link_status == link_status)
    items = query.order_by(MediaLink.created_at.desc()).all()

    # Pre-load media names
    media_map = {m.id: m for m in db.query(MediaContent).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Links"
    headers = ["ID", "Media Name", "Media Category", "Platform", "URL", "Description", "Status", "Link Category", "Created At"]
    style_header(ws, headers)

    for row, lnk in enumerate(items, 2):
        media = media_map.get(lnk.media_id)
        ws.cell(row=row, column=1, value=lnk.id)
        ws.cell(row=row, column=2, value=media.media_name if media else f"#{lnk.media_id}")
        ws.cell(row=row, column=3, value=media.media_category if media else "")
        ws.cell(row=row, column=4, value=lnk.platform)
        ws.cell(row=row, column=5, value=lnk.url)
        ws.cell(row=row, column=6, value=lnk.description or "")
        ws.cell(row=row, column=7, value=lnk.link_status or "active")
        ws.cell(row=row, column=8, value=lnk.link_category or "")
        ws.cell(row=row, column=9, value=str(lnk.created_at) if lnk.created_at else "")

    auto_width(ws)
    return workbook_to_response(wb, f"links_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


@router.get("/export/status", summary="Export all status tracking to Excel")
def export_status_excel(db: Session = Depends(get_db)):
    items = db.query(MediaStatus).order_by(MediaStatus.created_at.desc()).all()
    media_map = {m.id: m for m in db.query(MediaContent).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Status Tracking"
    headers = ["ID", "Media Name", "Media Category", "Status", "Notes", "Updated By", "Created At", "Updated At"]
    style_header(ws, headers)

    for row, s in enumerate(items, 2):
        media = media_map.get(s.media_id)
        ws.cell(row=row, column=1, value=s.id)
        ws.cell(row=row, column=2, value=media.media_name if media else f"#{s.media_id}")
        ws.cell(row=row, column=3, value=media.media_category if media else "")
        ws.cell(row=row, column=4, value=s.status)
        ws.cell(row=row, column=5, value=s.notes or "")
        ws.cell(row=row, column=6, value=s.updated_by or "")
        ws.cell(row=row, column=7, value=str(s.created_at) if s.created_at else "")
        ws.cell(row=row, column=8, value=str(s.updated_at) if s.updated_at else "")

    auto_width(ws)
    return workbook_to_response(wb, f"status_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ==================== TEMPLATES ====================

@router.get("/templates/media", summary="Download media import template")
def download_media_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Import"
    headers = ["media_category", "media_name", "release_date", "genre", "director", "cast_members", "rating", "review", "is_available", "available_on"]
    style_header(ws, headers)
    # Example rows
    examples = [
        ["movies", "The Dark Knight", "2008-07-18", "Action, Thriller", "Christopher Nolan", "Christian Bale, Heath Ledger", 9.0, "A masterpiece", "true", "Netflix, Amazon Prime"],
        ["webseries", "Breaking Bad", "2008-01-20", "Drama, Crime", "Vince Gilligan", "Bryan Cranston, Aaron Paul", 9.5, "Best TV series ever", "true", "Netflix"],
        ["music", "Bohemian Rhapsody", "1975-10-31", "Rock", "Queen", "Freddie Mercury", 9.0, "Iconic song", "false", ""],
    ]
    for row, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(italic=True, color="888888")
    auto_width(ws)
    return workbook_to_response(wb, "media_import_template.xlsx")


@router.get("/templates/links", summary="Download links import template")
def download_links_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Links Import"
    headers = ["media_name", "media_category", "platform", "url", "description", "link_status"]
    style_header(ws, headers)
    examples = [
        ["The Dark Knight", "movies", "youtube", "https://youtube.com/watch?v=abc123", "Official trailer", "active"],
        ["Breaking Bad", "webseries", "netflix", "https://netflix.com/title/123", "Full series", "active"],
        ["New Content", "", "instagram", "https://instagram.com/p/xyz", "Promo post", "active"],
    ]
    for row, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(italic=True, color="888888")
    # Add a note about media_category being optional
    ws.cell(row=6, column=1, value="Note: media_category is optional. If media doesn't exist, it will be auto-created (visible on Media tab).")
    ws.cell(row=6, column=1).font = Font(italic=True, color="FF6600", size=10)
    auto_width(ws)
    return workbook_to_response(wb, "links_import_template.xlsx")


@router.get("/templates/status", summary="Download status import template")
def download_status_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Status Import"
    headers = ["media_name", "status", "notes"]
    style_header(ws, headers)
    examples = [
        ["The Dark Knight", "completed", "Watched and reviewed"],
        ["Breaking Bad", "in_progress", "Currently on season 3"],
    ]
    for row, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(italic=True, color="888888")
    auto_width(ws)
    return workbook_to_response(wb, "status_import_template.xlsx")


# ==================== IMPORT ====================

BATCH_SIZE = 100  # Commit every N records for performance


def _persist_import_history(db: Session, import_type: str, file_name: str,
                            total_rows: int, total_processed: int,
                            successful: int, failed: int, skipped: int,
                            errors: list, success_records: list):
    """Save import run + individual record results to DB for history tracking."""
    success_rate = round((successful / total_processed * 100)) if total_processed > 0 else 0

    history = ImportHistory(
        import_type=import_type,
        file_name=file_name,
        total_rows=total_rows,
        total_processed=total_processed,
        successful=successful,
        failed=failed,
        skipped_empty=skipped,
        success_rate=success_rate,
    )
    db.add(history)
    db.flush()  # Get the ID

    # Store failed records
    for err in errors:
        rec = ImportRecord(
            history_id=history.id,
            row_number=err["row"],
            status="failed",
            record_name=err["data"].get("media_name") or err["data"].get("url") or "",
            record_data=err["data"],
            error_message=err["issue"],
            issue_type=err["issue_type"],
        )
        db.add(rec)

    # Store successful records (just name/row for reference)
    for sr in success_records:
        rec = ImportRecord(
            history_id=history.id,
            row_number=sr["row"],
            status="success",
            record_name=sr["name"],
            record_data=sr.get("data"),
        )
        db.add(rec)

    db.commit()
    return history.id


@router.post("/import/media", summary="Import media from Excel")
async def import_media_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    # --- Pre-load ALL existing media names in one query (fast bulk check) ---
    existing_names_rows = db.query(MediaContent.media_name).all()
    existing_names = {name.lower() for (name,) in existing_names_rows}

    # Track names seen within this file to catch intra-file duplicates
    seen_in_batch = set()

    created, skipped = 0, 0
    errors = []
    success_records = []
    pending_records = []

    for i, row in enumerate(rows, 2):
        try:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row[:2]):
                skipped += 1
                continue

            category = str(row[0]).strip().lower() if row[0] else ""
            name = str(row[1]).strip().title() if row[1] else ""
            row_data = {
                "media_category": category,
                "media_name": name,
                "release_date": str(row[2]) if len(row) > 2 and row[2] else "",
                "genre": str(row[3]).strip() if len(row) > 3 and row[3] else "",
                "director": str(row[4]).strip() if len(row) > 4 and row[4] else "",
                "rating": str(row[6]) if len(row) > 6 and row[6] else "",
            }

            if not category:
                errors.append({"row": i, "data": row_data, "issue": "media_category is required", "issue_type": "missing_field"})
                continue
            if not name:
                errors.append({"row": i, "data": row_data, "issue": "media_name is required", "issue_type": "missing_field"})
                continue

            name_lower = name.lower()

            # Check against intra-file duplicates (same name appearing multiple times in this sheet)
            if name_lower in seen_in_batch:
                errors.append({"row": i, "data": row_data, "issue": f"Duplicate within file: '{name}' already appears earlier in this sheet", "issue_type": "duplicate_in_file"})
                continue

            # Check against DB existing names
            if name_lower in existing_names:
                errors.append({"row": i, "data": row_data, "issue": f"Media '{name}' already exists in database", "issue_type": "duplicate"})
                continue

            # Validate rating range
            rating_val = None
            if len(row) > 6 and row[6] is not None and str(row[6]).strip() not in ("None", ""):
                try:
                    rating_val = float(row[6])
                    if rating_val < 0 or rating_val > 10:
                        errors.append({"row": i, "data": row_data, "issue": f"Rating {rating_val} is out of range (0-10)", "issue_type": "validation"})
                        continue
                except (ValueError, TypeError):
                    errors.append({"row": i, "data": row_data, "issue": f"Invalid rating value: '{row[6]}'", "issue_type": "validation"})
                    continue

            record = MediaContent(
                media_category=category,
                media_name=name,
                release_date=_parse_date(row[2] if len(row) > 2 else None),
                genre=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                director=str(row[4]).strip() if len(row) > 4 and row[4] else None,
                cast_members=str(row[5]).strip() if len(row) > 5 and row[5] else None,
                rating=rating_val,
                review=str(row[7]).strip() if len(row) > 7 and row[7] else None,
                is_available=str(row[8]).strip().lower() if len(row) > 8 and row[8] and str(row[8]).strip().lower() in ("true", "false") else "false",
                available_on=str(row[9]).strip() if len(row) > 9 and row[9] else None,
            )
            pending_records.append(record)
            seen_in_batch.add(name_lower)
            existing_names.add(name_lower)  # Prevent later rows from duplicating
            success_records.append({"row": i, "name": name, "data": row_data})
            created += 1

            # Commit in batches for performance
            if len(pending_records) >= BATCH_SIZE:
                db.add_all(pending_records)
                db.commit()
                pending_records = []

        except Exception as e:
            row_data = {"media_category": str(row[0]) if row and row[0] else "", "media_name": str(row[1]) if row and len(row) > 1 and row[1] else ""}
            errors.append({"row": i, "data": row_data, "issue": str(e), "issue_type": "exception"})

    # Commit remaining records
    if pending_records:
        db.add_all(pending_records)
        db.commit()

    total_processed = len(rows) - skipped

    # Persist import history
    history_id = _persist_import_history(
        db, "media", file.filename or "unknown.xlsx",
        len(rows), total_processed, created, len(errors), skipped,
        errors, success_records
    )

    return {
        "message": f"Imported {created} media items",
        "import_history_id": history_id,
        "summary": {
            "total_rows": len(rows),
            "total_processed": total_processed,
            "successful": created,
            "failed": len(errors),
            "skipped_empty": skipped,
        },
        "created": created,
        "errors": errors,
    }


@router.post("/import/links", summary="Import links from Excel")
async def import_links_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    # Build media name->id map (one query)
    all_media = db.query(MediaContent).all()
    media_map = {m.media_name.lower(): m for m in all_media}

    # Pre-load ALL existing URLs in one query
    existing_urls_rows = db.query(MediaLink.url).all()
    existing_urls = {url for (url,) in existing_urls_rows}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    created, skipped = 0, 0
    errors = []
    success_records = []
    pending_records = []
    seen_urls = set()

    for i, row in enumerate(rows, 2):
        try:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row[:4]):
                skipped += 1
                continue

            # Template columns: media_name, media_category, platform, url, description, link_status
            media_name_raw = str(row[0]).strip() if row[0] else ""
            media_category_raw = str(row[1]).strip().lower() if len(row) > 1 and row[1] and str(row[1]).strip() else ""
            platform_raw = str(row[2]).strip().lower() if len(row) > 2 and row[2] else "other"
            url_raw = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            description_raw = str(row[4]).strip() if len(row) > 4 and row[4] else None
            link_status_raw = str(row[5]).strip().lower() if len(row) > 5 and row[5] else "active"

            row_data = {
                "media_name": media_name_raw,
                "media_category": media_category_raw,
                "platform": platform_raw,
                "url": url_raw,
            }

            if not media_name_raw:
                errors.append({"row": i, "data": row_data, "issue": "media_name is required", "issue_type": "missing_field"})
                continue
            if not url_raw:
                errors.append({"row": i, "data": row_data, "issue": "URL is required", "issue_type": "missing_field"})
                continue

            media = media_map.get(media_name_raw.lower()) or media_map.get(media_name_raw.title().lower())

            # Auto-create media entry if not found
            if not media:
                category_to_use = media_category_raw if media_category_raw else "uncategorized"
                new_media = MediaContent(
                    media_category=category_to_use,
                    media_name=media_name_raw.title(),
                )
                db.add(new_media)
                db.flush()  # Get ID without full commit
                media = new_media
                media_map[media_name_raw.lower()] = media  # Cache for next rows

            # Check against DB existing URLs
            if url_raw in existing_urls:
                errors.append({"row": i, "data": row_data, "issue": "Link URL already exists in database", "issue_type": "duplicate"})
                continue

            # Check intra-file duplicate URLs
            if url_raw in seen_urls:
                errors.append({"row": i, "data": row_data, "issue": "Duplicate URL within this file", "issue_type": "duplicate_in_file"})
                continue

            record = MediaLink(
                media_id=media.id,
                platform=platform_raw,
                url=url_raw,
                description=description_raw,
                link_status=link_status_raw,
                link_category=media.media_category,
            )
            pending_records.append(record)
            seen_urls.add(url_raw)
            existing_urls.add(url_raw)
            success_records.append({"row": i, "name": url_raw, "data": row_data})
            created += 1

            if len(pending_records) >= BATCH_SIZE:
                db.add_all(pending_records)
                db.commit()
                pending_records = []

        except Exception as e:
            row_data = {"media_name": str(row[0]) if row and row[0] else "", "url": str(row[3]) if row and len(row) > 3 and row[3] else ""}
            errors.append({"row": i, "data": row_data, "issue": str(e), "issue_type": "exception"})

    if pending_records:
        db.add_all(pending_records)
        db.commit()

    total_processed = len(rows) - skipped
    history_id = _persist_import_history(
        db, "links", file.filename or "unknown.xlsx",
        len(rows), total_processed, created, len(errors), skipped,
        errors, success_records
    )
    return {
        "message": f"Imported {created} links",
        "import_history_id": history_id,
        "summary": {
            "total_rows": len(rows),
            "total_processed": total_processed,
            "successful": created,
            "failed": len(errors),
            "skipped_empty": skipped,
        },
        "created": created,
        "errors": errors,
    }


@router.post("/import/status", summary="Import status tracking from Excel")
async def import_status_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    all_media = db.query(MediaContent).all()
    media_map = {m.media_name.lower(): m for m in all_media}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    created, skipped = 0, 0
    errors = []
    success_records = []
    pending_records = []

    for i, row in enumerate(rows, 2):
        try:
            if not row or all(cell is None or str(cell).strip() == "" for cell in row[:2]):
                skipped += 1
                continue

            media_name_raw = str(row[0]).strip() if row[0] else ""
            status_raw = str(row[1]).strip().lower() if len(row) > 1 and row[1] else ""
            row_data = {
                "media_name": media_name_raw,
                "status": status_raw,
                "notes": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            }

            if not media_name_raw:
                errors.append({"row": i, "data": row_data, "issue": "media_name is required", "issue_type": "missing_field"})
                continue
            if not status_raw:
                errors.append({"row": i, "data": row_data, "issue": "status is required", "issue_type": "missing_field"})
                continue

            media = media_map.get(media_name_raw.lower()) or media_map.get(media_name_raw.title().lower())
            if not media:
                errors.append({"row": i, "data": row_data, "issue": f"Media '{media_name_raw}' not found. Create it first.", "issue_type": "reference_missing"})
                continue

            record = MediaStatus(
                media_id=media.id,
                status=status_raw,
                notes=str(row[2]).strip() if len(row) > 2 and row[2] else None,
            )
            pending_records.append(record)
            success_records.append({"row": i, "name": media_name_raw, "data": row_data})
            created += 1

            if len(pending_records) >= BATCH_SIZE:
                db.add_all(pending_records)
                db.commit()
                pending_records = []

        except Exception as e:
            row_data = {"media_name": str(row[0]) if row and row[0] else "", "status": str(row[1]) if row and len(row) > 1 and row[1] else ""}
            errors.append({"row": i, "data": row_data, "issue": str(e), "issue_type": "exception"})

    if pending_records:
        db.add_all(pending_records)
        db.commit()

    total_processed = len(rows) - skipped
    history_id = _persist_import_history(
        db, "status", file.filename or "unknown.xlsx",
        len(rows), total_processed, created, len(errors), skipped,
        errors, success_records
    )
    return {
        "message": f"Imported {created} status entries",
        "import_history_id": history_id,
        "summary": {
            "total_rows": len(rows),
            "total_processed": total_processed,
            "successful": created,
            "failed": len(errors),
            "skipped_empty": skipped,
        },
        "created": created,
        "errors": errors,
    }


# ==================== FULL EXPORT (Category + Tags tabs) ====================

@router.get("/export/full", summary="Export all data with category tabs and tags sheet")
def export_full_excel(db: Session = Depends(get_db)):
    """Export all media in a workbook with separate sheets per category + a Tags summary sheet."""
    all_media = db.query(MediaContent).order_by(MediaContent.updated_at.desc()).all()

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group by category
    categories = {}
    for m in all_media:
        categories.setdefault(m.media_category, []).append(m)

    headers = ["ID", "Name", "Release Date", "Genre", "Director", "Cast", "Rating", "Review", "Tags", "Is Available", "Available On", "Updated At"]

    for cat, items in categories.items():
        ws = wb.create_sheet(title=cat.capitalize()[:31])
        style_header(ws, headers)
        for row, m in enumerate(items, 2):
            ws.cell(row=row, column=1, value=m.id)
            ws.cell(row=row, column=2, value=m.media_name)
            ws.cell(row=row, column=3, value=str(m.release_date) if m.release_date else "")
            ws.cell(row=row, column=4, value=m.genre or "")
            ws.cell(row=row, column=5, value=m.director or "")
            ws.cell(row=row, column=6, value=m.cast_members or "")
            ws.cell(row=row, column=7, value=m.rating)
            ws.cell(row=row, column=8, value=m.review or "")
            ws.cell(row=row, column=9, value=m.tags or "")
            ws.cell(row=row, column=10, value=m.is_available or "false")
            ws.cell(row=row, column=11, value=m.available_on or "")
            ws.cell(row=row, column=12, value=str(m.updated_at) if m.updated_at else "")
        auto_width(ws)

    # Tags summary sheet
    tag_counts = {}
    for m in all_media:
        if m.tags:
            for tag in m.tags.split(","):
                tag = tag.strip()
                if tag:
                    tag_counts[tag] = tag_counts.get(tag, 0) + 1

    ws_tags = wb.create_sheet(title="Tags")
    tag_headers = ["Tag", "Count", "Media Names"]
    style_header(ws_tags, tag_headers)

    # Build tag -> media names mapping
    tag_media = {}
    for m in all_media:
        if m.tags:
            for tag in m.tags.split(","):
                tag = tag.strip()
                if tag:
                    tag_media.setdefault(tag, []).append(m.media_name)

    sorted_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)
    for row, (tag, count) in enumerate(sorted_tags, 2):
        ws_tags.cell(row=row, column=1, value=tag)
        ws_tags.cell(row=row, column=2, value=count)
        ws_tags.cell(row=row, column=3, value=", ".join(tag_media.get(tag, [])))
    auto_width(ws_tags)

    if not wb.sheetnames:
        wb.create_sheet(title="Empty")

    return workbook_to_response(wb, f"full_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ==================== IMPORT HISTORY ====================

@router.get("/import-history", summary="List import history")
def list_import_history(
    import_type: str | None = Query(None),
    limit: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """List all past import runs, most recent first."""
    query = db.query(ImportHistory)
    if import_type:
        query = query.filter(ImportHistory.import_type == import_type)
    items = query.order_by(desc(ImportHistory.created_at)).limit(limit).all()
    return [
        {
            "id": h.id,
            "import_type": h.import_type,
            "file_name": h.file_name,
            "total_rows": h.total_rows,
            "total_processed": h.total_processed,
            "successful": h.successful,
            "failed": h.failed,
            "skipped_empty": h.skipped_empty,
            "success_rate": h.success_rate,
            "created_at": h.created_at.isoformat() if h.created_at else None,
        }
        for h in items
    ]


@router.get("/import-history/{history_id}/records", summary="Get records for a specific import")
def get_import_records(
    history_id: int,
    status: str | None = Query(None, description="Filter: success, failed, skipped"),
    search: str | None = Query(None, description="Search record names"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
    db: Session = Depends(get_db),
):
    """Get individual record results for a specific import run with filter/search."""
    query = db.query(ImportRecord).filter(ImportRecord.history_id == history_id)
    if status:
        query = query.filter(ImportRecord.status == status)
    if search:
        query = query.filter(ImportRecord.record_name.ilike(f"%{search}%"))

    total = query.count()
    items = query.order_by(ImportRecord.row_number).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "records": [
            {
                "id": r.id,
                "row_number": r.row_number,
                "status": r.status,
                "record_name": r.record_name,
                "record_data": r.record_data,
                "error_message": r.error_message,
                "issue_type": r.issue_type,
            }
            for r in items
        ],
    }


@router.get("/import-history/{history_id}/export", summary="Export import records as CSV")
def export_import_records(
    history_id: int,
    status: str | None = Query(None, description="Filter: success, failed"),
    format: str = Query("csv", description="Export format: csv or excel"),
    db: Session = Depends(get_db),
):
    """Export records from a specific import as CSV or Excel for user to fix and re-upload."""
    history = db.query(ImportHistory).filter(ImportHistory.id == history_id).first()
    if not history:
        raise HTTPException(status_code=404, detail="Import history not found")

    query = db.query(ImportRecord).filter(ImportRecord.history_id == history_id)
    if status:
        query = query.filter(ImportRecord.status == status)
    records = query.order_by(ImportRecord.row_number).all()

    if format == "excel":
        wb = Workbook()
        ws = wb.active
        ws.title = "Import Records"
        headers = ["Row #", "Status", "Name", "Issue Type", "Error Message", "Data"]
        style_header(ws, headers)
        for idx, r in enumerate(records, 2):
            ws.cell(row=idx, column=1, value=r.row_number)
            ws.cell(row=idx, column=2, value=r.status)
            ws.cell(row=idx, column=3, value=r.record_name or "")
            ws.cell(row=idx, column=4, value=r.issue_type or "")
            ws.cell(row=idx, column=5, value=r.error_message or "")
            ws.cell(row=idx, column=6, value=str(r.record_data) if r.record_data else "")
            # Highlight failed rows in red
            if r.status == "failed":
                for col in range(1, 7):
                    ws.cell(row=idx, column=col).fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        auto_width(ws)
        return workbook_to_response(wb, f"import_records_{history_id}_{status or 'all'}.xlsx")
    else:
        # CSV export
        buf = io.StringIO()
        writer = csv.writer(buf)
        writer.writerow(["Row #", "Status", "Name", "Issue Type", "Error Message", "Data"])
        for r in records:
            writer.writerow([
                r.row_number, r.status, r.record_name or "",
                r.issue_type or "", r.error_message or "",
                str(r.record_data) if r.record_data else "",
            ])
        buf.seek(0)
        return StreamingResponse(
            io.BytesIO(buf.getvalue().encode("utf-8")),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=import_records_{history_id}_{status or 'all'}.csv"},
        )


# ==================== IMPORT SCHEDULE ====================

@router.get("/schedule", summary="Get import schedules")
def get_schedules(db: Session = Depends(get_db)):
    items = db.query(ImportSchedule).order_by(ImportSchedule.import_type).all()
    return [
        {
            "id": s.id,
            "import_type": s.import_type,
            "interval": s.interval,
            "is_active": s.is_active,
            "last_run": s.last_run.isoformat() if s.last_run else None,
            "next_run": s.next_run.isoformat() if s.next_run else None,
            "notify_email": s.notify_email,
            "notify_in_app": s.notify_in_app,
        }
        for s in items
    ]


@router.post("/schedule", summary="Create or update import schedule")
def save_schedule(
    import_type: str = Query(...),
    interval: str = Query(..., description="daily, weekly, monthly"),
    notify_email: str | None = Query(None),
    notify_in_app: bool = Query(True),
    db: Session = Depends(get_db),
):
    """Save or update a schedule config for auto-imports."""
    if interval not in ("daily", "weekly", "monthly"):
        raise HTTPException(status_code=400, detail="Interval must be daily, weekly, or monthly")

    existing = db.query(ImportSchedule).filter(ImportSchedule.import_type == import_type).first()
    now = datetime.now()
    delta = {"daily": timedelta(days=1), "weekly": timedelta(weeks=1), "monthly": timedelta(days=30)}
    next_run = now + delta[interval]

    if existing:
        existing.interval = interval
        existing.is_active = True
        existing.notify_email = notify_email
        existing.notify_in_app = notify_in_app
        existing.next_run = next_run
    else:
        schedule = ImportSchedule(
            import_type=import_type,
            interval=interval,
            is_active=True,
            next_run=next_run,
            notify_email=notify_email,
            notify_in_app=notify_in_app,
        )
        db.add(schedule)

    db.commit()
    return {"message": f"Schedule for {import_type} set to {interval}", "next_run": next_run.isoformat()}


@router.delete("/schedule/{schedule_id}", summary="Delete import schedule")
def delete_schedule(schedule_id: int, db: Session = Depends(get_db)):
    schedule = db.query(ImportSchedule).filter(ImportSchedule.id == schedule_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Schedule not found")
    db.delete(schedule)
    db.commit()
    return {"message": "Schedule deleted"}

