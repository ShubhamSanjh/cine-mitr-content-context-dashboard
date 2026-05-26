"""
Excel Export, Import, and Template endpoints.
Supports: Media, Links, Status Tracking data.
Uses openpyxl for .xlsx generation and parsing.
"""

import io
from datetime import datetime, date as date_type
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.models.media import MediaContent, MediaLink, MediaStatus

router = APIRouter(prefix="/excel")


def _parse_date(val):
    """Convert various date formats from Excel to Python date object."""
    if val is None:
        return None
    if isinstance(val, date_type):
        return val
    if isinstance(val, datetime):
        return val.date()
    s = str(val).strip()
    if not s or s in ("None", ""):
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None

# --- Styling helpers ---
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, headers):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)


def workbook_to_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# ==================== EXPORT ====================

@router.get("/export/media", summary="Export all media to Excel")
def export_media_excel(
    media_category: str | None = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(MediaContent)
    if media_category:
        query = query.filter(MediaContent.media_category == media_category)
    items = query.order_by(MediaContent.updated_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Media"
    headers = ["ID", "Category", "Name", "Release Date", "Genre", "Director", "Cast", "Rating", "Review", "Tags", "Is Available", "Available On", "Created At"]
    style_header(ws, headers)

    for row, m in enumerate(items, 2):
        ws.cell(row=row, column=1, value=m.id)
        ws.cell(row=row, column=2, value=m.media_category)
        ws.cell(row=row, column=3, value=m.media_name)
        ws.cell(row=row, column=4, value=str(m.release_date) if m.release_date else "")
        ws.cell(row=row, column=5, value=m.genre or "")
        ws.cell(row=row, column=6, value=m.director or "")
        ws.cell(row=row, column=7, value=m.cast_members or "")
        ws.cell(row=row, column=8, value=m.rating)
        ws.cell(row=row, column=9, value=m.review or "")
        ws.cell(row=row, column=10, value=m.tags or "")
        ws.cell(row=row, column=11, value=m.is_available or "false")
        ws.cell(row=row, column=12, value=m.available_on or "")
        ws.cell(row=row, column=13, value=str(m.created_at) if m.created_at else "")

    auto_width(ws)
    return workbook_to_response(wb, f"media_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


@router.get("/export/links", summary="Export all links to Excel")
def export_links_excel(
    link_category: str | None = Query(None),
    link_status: str | None = Query(None),
    db: Session = Depends(get_db),
):
    query = db.query(MediaLink)
    if link_category:
        query = query.filter(MediaLink.link_category == link_category)
    if link_status:
        query = query.filter(MediaLink.link_status == link_status)
    items = query.order_by(MediaLink.created_at.desc()).all()

    # Pre-load media names
    media_map = {m.id: m for m in db.query(MediaContent).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Links"
    headers = ["ID", "Media Name", "Media Category", "Platform", "URL", "Description", "Status", "Category", "Created At"]
    style_header(ws, headers)

    for row, lnk in enumerate(items, 2):
        media = media_map.get(lnk.media_id)
        ws.cell(row=row, column=1, value=lnk.id)
        ws.cell(row=row, column=2, value=media.media_name if media else f"#{lnk.media_id}")
        ws.cell(row=row, column=3, value=media.media_category if media else "")
        ws.cell(row=row, column=4, value=lnk.platform)
        ws.cell(row=row, column=5, value=lnk.url)
        ws.cell(row=row, column=6, value=lnk.description or "")
        ws.cell(row=row, column=7, value=lnk.link_status or "active")
        ws.cell(row=row, column=8, value=lnk.link_category or "")
        ws.cell(row=row, column=9, value=str(lnk.created_at) if lnk.created_at else "")

    auto_width(ws)
    return workbook_to_response(wb, f"links_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


@router.get("/export/status", summary="Export all status tracking to Excel")
def export_status_excel(db: Session = Depends(get_db)):
    items = db.query(MediaStatus).order_by(MediaStatus.created_at.desc()).all()
    media_map = {m.id: m for m in db.query(MediaContent).all()}

    wb = Workbook()
    ws = wb.active
    ws.title = "Status Tracking"
    headers = ["ID", "Media Name", "Media Category", "Status", "Notes", "Updated By", "Created At", "Updated At"]
    style_header(ws, headers)

    for row, s in enumerate(items, 2):
        media = media_map.get(s.media_id)
        ws.cell(row=row, column=1, value=s.id)
        ws.cell(row=row, column=2, value=media.media_name if media else f"#{s.media_id}")
        ws.cell(row=row, column=3, value=media.media_category if media else "")
        ws.cell(row=row, column=4, value=s.status)
        ws.cell(row=row, column=5, value=s.notes or "")
        ws.cell(row=row, column=6, value=s.updated_by or "")
        ws.cell(row=row, column=7, value=str(s.created_at) if s.created_at else "")
        ws.cell(row=row, column=8, value=str(s.updated_at) if s.updated_at else "")

    auto_width(ws)
    return workbook_to_response(wb, f"status_export_{datetime.now().strftime('%Y%m%d')}.xlsx")


# ==================== TEMPLATES ====================

@router.get("/templates/media", summary="Download media import template")
def download_media_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Media Import"
    headers = ["media_category", "media_name", "release_date", "genre", "director", "cast_members", "rating", "review", "is_available", "available_on"]
    style_header(ws, headers)
    # Example rows
    examples = [
        ["movies", "The Dark Knight", "2008-07-18", "Action, Thriller", "Christopher Nolan", "Christian Bale, Heath Ledger", 9.0, "A masterpiece", "true", "Netflix, Amazon Prime"],
        ["webseries", "Breaking Bad", "2008-01-20", "Drama, Crime", "Vince Gilligan", "Bryan Cranston, Aaron Paul", 9.5, "Best TV series ever", "true", "Netflix"],
        ["music", "Bohemian Rhapsody", "1975-10-31", "Rock", "Queen", "Freddie Mercury", 9.0, "Iconic song", "false", ""],
    ]
    for row, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(italic=True, color="888888")
    auto_width(ws)
    return workbook_to_response(wb, "media_import_template.xlsx")


@router.get("/templates/links", summary="Download links import template")
def download_links_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Links Import"
    headers = ["media_name", "platform", "url", "description", "link_status"]
    style_header(ws, headers)
    examples = [
        ["The Dark Knight", "youtube", "https://youtube.com/watch?v=abc123", "Official trailer", "active"],
        ["Breaking Bad", "netflix", "https://netflix.com/title/123", "Full series", "active"],
    ]
    for row, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(italic=True, color="888888")
    auto_width(ws)
    return workbook_to_response(wb, "links_import_template.xlsx")


@router.get("/templates/status", summary="Download status import template")
def download_status_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Status Import"
    headers = ["media_name", "status", "notes"]
    style_header(ws, headers)
    examples = [
        ["The Dark Knight", "completed", "Watched and reviewed"],
        ["Breaking Bad", "in_progress", "Currently on season 3"],
    ]
    for row, ex in enumerate(examples, 2):
        for col, val in enumerate(ex, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.font = Font(italic=True, color="888888")
    auto_width(ws)
    return workbook_to_response(wb, "status_import_template.xlsx")


# ==================== IMPORT ====================

@router.post("/import/media", summary="Import media from Excel")
async def import_media_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created, errors = 0, []

    for i, row in enumerate(rows, 2):
        try:
            if not row or not row[0] or not row[1]:
                continue  # skip empty rows
            category = str(row[0]).strip().lower()
            name = str(row[1]).strip().title()
            if not name:
                errors.append(f"Row {i}: media_name is required")
                continue

            # Skip duplicates by name (case-insensitive)
            existing = db.query(MediaContent).filter(
                MediaContent.media_name.ilike(name)
            ).first()
            if existing:
                errors.append(f"Row {i}: Media '{name}' already exists (id={existing.id}), skipped")
                continue

            record = MediaContent(
                media_category=category,
                media_name=name,
                release_date=_parse_date(row[2]),
                genre=str(row[3]).strip() if row[3] else None,
                director=str(row[4]).strip() if row[4] else None,
                cast_members=str(row[5]).strip() if row[5] else None,
                rating=float(row[6]) if row[6] is not None and str(row[6]).strip() not in ("None", "") else None,
                review=str(row[7]).strip() if len(row) > 7 and row[7] else None,
                is_available=str(row[8]).strip().lower() if len(row) > 8 and row[8] and str(row[8]).strip().lower() in ("true", "false") else "false",
                available_on=str(row[9]).strip() if len(row) > 9 and row[9] else None,
            )
            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {"message": f"Imported {created} media items", "created": created, "errors": errors}


@router.post("/import/links", summary="Import links from Excel")
async def import_links_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    # Build media name->id map
    all_media = db.query(MediaContent).all()
    media_map = {m.media_name.lower(): m for m in all_media}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created, errors = 0, []

    for i, row in enumerate(rows, 2):
        try:
            if not row or not row[0] or not row[2]:
                continue
            media_name = str(row[0]).strip().lower()
            media = media_map.get(media_name)
            if not media:
                # Try title-case match
                media = media_map.get(media_name.title().lower())
            if not media:
                errors.append(f"Row {i}: Media '{row[0]}' not found. Create it first.")
                continue

            record = MediaLink(
                media_id=media.id,
                platform=str(row[1]).strip().lower() if row[1] else "other",
                url=str(row[2]).strip(),
                description=str(row[3]).strip() if len(row) > 3 and row[3] else None,
                link_status=str(row[4]).strip().lower() if len(row) > 4 and row[4] else "active",
                link_category=media.media_category,
            )

            # Skip duplicate URLs
            existing_link = db.query(MediaLink).filter(
                MediaLink.url == record.url
            ).first()
            if existing_link:
                errors.append(f"Row {i}: Link URL already exists (link_id={existing_link.id}), skipped")
                continue

            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {"message": f"Imported {created} links", "created": created, "errors": errors}


@router.post("/import/status", summary="Import status tracking from Excel")
async def import_status_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx files are accepted")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active

    all_media = db.query(MediaContent).all()
    media_map = {m.media_name.lower(): m for m in all_media}

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    created, errors = 0, []

    for i, row in enumerate(rows, 2):
        try:
            if not row or not row[0] or not row[1]:
                continue
            media_name = str(row[0]).strip().lower()
            media = media_map.get(media_name)
            if not media:
                media = media_map.get(media_name.title().lower())
            if not media:
                errors.append(f"Row {i}: Media '{row[0]}' not found. Create it first.")
                continue

            record = MediaStatus(
                media_id=media.id,
                status=str(row[1]).strip().lower(),
                notes=str(row[2]).strip() if len(row) > 2 and row[2] else None,
            )
            db.add(record)
            created += 1
        except Exception as e:
            errors.append(f"Row {i}: {str(e)}")

    db.commit()
    return {"message": f"Imported {created} status entries", "created": created, "errors": errors}


# ==================== FULL EXPORT (Category + Tags tabs) ====================

@router.get("/export/full", summary="Export all data with category tabs and tags sheet")
def export_full_excel(db: Session = Depends(get_db)):
    """Export all media in a workbook with separate sheets per category + a Tags summary sheet."""
    all_media = db.query(MediaContent).order_by(MediaContent.updated_at.desc()).all()

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Group by category
    categories = {}
    for m in all_media:
        categories.setdefault(m.media_category, []).append(m)

    headers = ["ID", "Name", "Release Date", "Genre", "Director", "Cast", "Rating", "Review", "Tags", "Is Available", "Available On", "Updated At"]

    for cat, items in categories.items():
        ws = wb.create_sheet(title=cat.capitalize()[:31])
        style_header(ws, headers)
        for row, m in enumerate(items, 2):
            ws.cell(row=row, column=1, value=m.id)
            ws.cell(row=row, column=2, value=m.media_name)
            ws.cell(row=row, column=3, value=str(m.release_date) if m.release_date else "")
            ws.cell(row=row, column=4, value=m.genre or "")
            ws.cell(row=row, column=5, value=m.director or "")
            ws.cell(row=row, column=6, value=m.cast_members or "")
            ws.cell(row=row, column=7, value=m.rating)
            ws.cell(row=row, column=8, value=m.review or "")
            ws.cell(row=row, column=9, value=m.tags or "")
            ws.cell(row=row, column=10, value=m.is_available or "false")
            ws.cell(row=row, column=11, value=m.available_on or "")
            ws.cell(row=row, column=12, value=str(m.updated_at) if m.updated_at else "")
        auto_width(ws)

    # Tags summary sheet
    tag_counts = {}
    for m in all_media:
        if m.tags:
            for tag in m.tags.split(","):
                tag = tag.strip()
                if tag:
                    tag_counts[tag] = tag_counts.get(tag, 0) + 1

    ws_tags = wb.create_sheet(title="Tags")
    tag_headers = ["Tag", "Count", "Media Names"]
    style_header(ws_tags, tag_headers)

    # Build tag -> media names mapping
    tag_media = {}
    for m in all_media:
        if m.tags:
            for tag in m.tags.split(","):
                tag = tag.strip()
                if tag:
                    tag_media.setdefault(tag, []).append(m.media_name)

    sorted_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)
    for row, (tag, count) in enumerate(sorted_tags, 2):
        ws_tags.cell(row=row, column=1, value=tag)
        ws_tags.cell(row=row, column=2, value=count)
        ws_tags.cell(row=row, column=3, value=", ".join(tag_media.get(tag, [])))
    auto_width(ws_tags)

    if not wb.sheetnames:
        wb.create_sheet(title="Empty")

    return workbook_to_response(wb, f"full_export_{datetime.now().strftime('%Y%m%d')}.xlsx")

