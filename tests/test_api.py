"""
Basic tests for the Content Dashboard API.
Run: pytest tests/ -v
"""

import pytest
from fastapi.testclient import TestClient
from app.database import Base, engine
from app.main import app


@pytest.fixture(autouse=True)
def setup_db():
    """Ensure tables exist before each test."""
    Base.metadata.create_all(bind=engine)
    yield
    Base.metadata.drop_all(bind=engine)


client = TestClient(app)


def test_health():
    res = client.get("/api/v1/health")
    assert res.status_code == 200
    data = res.json()
    assert data["status"] == "healthy"


def test_list_contents_empty():
    res = client.get("/api/v1/contents/")
    assert res.status_code == 200
    data = res.json()
    assert data["total"] >= 0


def test_create_content():
    payload = {
        "title": "Test Item",
        "category": "reports",
        "description": "A test content record.",
        "status": "active",
        "is_published": False,
    }
    res = client.post("/api/v1/contents/", json=payload)
    assert res.status_code == 201
    data = res.json()
    assert data["title"] == "Test Item"
    assert data["id"] > 0


def test_get_content():
    payload = {"title": "Get Test", "category": "plans"}
    create_res = client.post("/api/v1/contents/", json=payload)
    cid = create_res.json()["id"]

    res = client.get(f"/api/v1/contents/{cid}")
    assert res.status_code == 200
    assert res.json()["title"] == "Get Test"


def test_update_content():
    payload = {"title": "Update Test", "category": "analytics"}
    create_res = client.post("/api/v1/contents/", json=payload)
    cid = create_res.json()["id"]

    update_res = client.put(f"/api/v1/contents/{cid}", json={"title": "Updated Title"})
    assert update_res.status_code == 200
    assert update_res.json()["title"] == "Updated Title"


def test_delete_content():
    payload = {"title": "Delete Me", "category": "documentation"}
    create_res = client.post("/api/v1/contents/", json=payload)
    cid = create_res.json()["id"]

    del_res = client.delete(f"/api/v1/contents/{cid}")
    assert del_res.status_code == 204

    get_res = client.get(f"/api/v1/contents/{cid}")
    assert get_res.status_code == 404


def test_seed_data():
    res = client.post("/api/v1/contents/seed")
    assert res.status_code == 200
    assert "inserted" in res.json()["message"]


def test_serve_ui():
    res = client.get("/")
    assert res.status_code == 200
    assert "Content Dashboard" in res.text


# ===== STATUS DEFINITIONS TESTS =====

def test_status_definitions_empty():
    res = client.get("/api/v1/status-definitions/")
    assert res.status_code == 200
    assert isinstance(res.json(), list)


def test_seed_default_statuses():
    res = client.post("/api/v1/status-definitions/seed")
    assert res.status_code == 200
    data = res.json()
    assert data["total"] > 0
    assert data["message"].startswith("Seeded")


def test_seed_default_statuses_idempotent():
    client.post("/api/v1/status-definitions/seed")
    res = client.post("/api/v1/status-definitions/seed")
    assert res.status_code == 200
    assert "Seeded 0" in res.json()["message"]


def test_create_status_definition():
    payload = {"name": "test_status", "label": "Test Status", "usage_context": "all"}
    res = client.post("/api/v1/status-definitions/", json=payload)
    assert res.status_code == 201
    data = res.json()
    assert data["name"] == "test_status"
    assert data["label"] == "Test Status"
    assert data["is_active"] == "true"


def test_create_duplicate_status_definition():
    payload = {"name": "duplicate", "label": "Duplicate"}
    client.post("/api/v1/status-definitions/", json=payload)
    res = client.post("/api/v1/status-definitions/", json=payload)
    assert res.status_code == 409


def test_update_status_definition():
    payload = {"name": "updatable", "label": "Updatable"}
    create_res = client.post("/api/v1/status-definitions/", json=payload)
    def_id = create_res.json()["id"]

    res = client.put(f"/api/v1/status-definitions/{def_id}", json={"label": "Updated Label"})
    assert res.status_code == 200
    assert res.json()["label"] == "Updated Label"


def test_delete_status_definition():
    payload = {"name": "deletable", "label": "Deletable"}
    create_res = client.post("/api/v1/status-definitions/", json=payload)
    def_id = create_res.json()["id"]

    res = client.delete(f"/api/v1/status-definitions/{def_id}")
    assert res.status_code == 204


def test_list_status_definitions_by_context():
    client.post("/api/v1/status-definitions/seed")
    res = client.get("/api/v1/status-definitions/?context=status_tracking")
    assert res.status_code == 200
    defs = res.json()
    for d in defs:
        assert d["usage_context"] in ("status_tracking", "all")


# ===== MEDIA TESTS =====

def test_create_media():
    payload = {"media_category": "movies", "media_name": "Test Movie"}
    res = client.post("/api/v1/media/", json=payload)
    assert res.status_code == 201
    assert res.json()["media_name"] == "Test Movie"


def test_list_media():
    client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Movie 1"})
    client.post("/api/v1/media/", json={"media_category": "shows", "media_name": "Show 1"})
    res = client.get("/api/v1/media/?page_size=50")
    assert res.status_code == 200
    assert res.json()["total"] >= 2


def test_update_media():
    create_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Old Name"})
    mid = create_res.json()["id"]
    res = client.put(f"/api/v1/media/{mid}", json={"media_name": "New Name"})
    assert res.status_code == 200
    assert res.json()["media_name"] == "New Name"


def test_delete_media():
    create_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Delete Me"})
    mid = create_res.json()["id"]
    res = client.delete(f"/api/v1/media/{mid}")
    assert res.status_code == 204


# ===== STATUS TRACKING WITH DYNAMIC VALIDATION =====

def test_create_status_with_seeded_definitions():
    """Status tracking should accept statuses from seeded status definitions."""
    client.post("/api/v1/status-definitions/seed")
    media_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Status Test"})
    media_id = media_res.json()["id"]

    res = client.post("/api/v1/media-status/", json={"media_id": media_id, "status": "in_progress"})
    assert res.status_code == 201
    assert res.json()["status"] == "in_progress"


def test_create_status_with_new_custom_definition():
    """After adding a custom status definition, it should be accepted."""
    client.post("/api/v1/status-definitions/", json={"name": "my_custom", "label": "My Custom", "usage_context": "status_tracking"})
    media_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Custom Status Test"})
    media_id = media_res.json()["id"]

    res = client.post("/api/v1/media-status/", json={"media_id": media_id, "status": "my_custom"})
    assert res.status_code == 201
    assert res.json()["status"] == "my_custom"


def test_create_status_invalid():
    """Invalid statuses should be rejected."""
    client.post("/api/v1/status-definitions/seed")
    media_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Invalid Status Test"})
    media_id = media_res.json()["id"]

    res = client.post("/api/v1/media-status/", json={"media_id": media_id, "status": "nonexistent_status"})
    assert res.status_code == 400


def test_valid_statuses_endpoint_dynamic():
    """The valid-statuses endpoint should return dynamic statuses from definitions."""
    client.post("/api/v1/status-definitions/seed")
    res = client.get("/api/v1/media-status/valid-statuses")
    assert res.status_code == 200
    statuses = res.json()["statuses"]
    assert "in_progress" in statuses
    assert "reviewed" in statuses


def test_status_tracking_list():
    client.post("/api/v1/status-definitions/seed")
    media_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "List Test"})
    media_id = media_res.json()["id"]
    client.post("/api/v1/media-status/", json={"media_id": media_id, "status": "planned"})

    res = client.get("/api/v1/media-status/")
    assert res.status_code == 200
    assert len(res.json()) >= 1


def test_update_status():
    client.post("/api/v1/status-definitions/seed")
    media_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Update Status Test"})
    media_id = media_res.json()["id"]
    status_res = client.post("/api/v1/media-status/", json={"media_id": media_id, "status": "planned"})
    status_id = status_res.json()["id"]

    res = client.put(f"/api/v1/media-status/{status_id}", json={"status": "completed", "notes": "Done!"})
    assert res.status_code == 200
    assert res.json()["status"] == "completed"
    assert res.json()["notes"] == "Done!"


def test_delete_status():
    client.post("/api/v1/status-definitions/seed")
    media_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Delete Status Test"})
    media_id = media_res.json()["id"]
    status_res = client.post("/api/v1/media-status/", json={"media_id": media_id, "status": "planned"})
    status_id = status_res.json()["id"]

    res = client.delete(f"/api/v1/media-status/{status_id}")
    assert res.status_code == 204


# ===== UI HTML STRUCTURE TESTS =====

def test_ui_has_dynamic_status_dropdown():
    """UI should have dynamically populated status dropdown."""
    res = client.get("/")
    assert 'id="fStatusValue"' in res.text


def test_ui_version():
    res = client.get("/")
    assert "v4.0.0" in res.text


def test_ui_analytics_page_exists():
    res = client.get("/")
    assert 'id="page-analytics"' in res.text
    assert 'id="analyticsTagChart"' in res.text


def test_ui_status_mgmt_page_exists():
    res = client.get("/")
    assert 'id="page-status-mgmt"' in res.text
    assert 'id="statusDefsTable"' in res.text


# ===== DASHBOARD TESTS =====

def test_dashboard_stats():
    res = client.get("/api/v1/dashboard/stats")
    assert res.status_code == 200
    data = res.json()
    assert "total_media" in data
    assert "total_links" in data


# ===== FEATURE 64: DASHBOARD FILTERS =====

def test_dashboard_stats_with_category_filter():
    client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Filter Test"})
    client.post("/api/v1/media/", json={"media_category": "shows", "media_name": "Show Filter"})
    res = client.get("/api/v1/dashboard/stats?media_category=movies")
    assert res.status_code == 200
    data = res.json()
    assert data["total_movies"] >= 1
    assert data["total_shows"] == 0


# ===== FEATURE 67: MEDIA NAME TITLE CASE =====

def test_media_name_title_case_on_create():
    res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "the dark knight"})
    assert res.status_code == 201
    assert res.json()["media_name"] == "The Dark Knight"


def test_media_name_title_case_on_update():
    create_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "test"})
    mid = create_res.json()["id"]
    res = client.put(f"/api/v1/media/{mid}", json={"media_name": "breaking bad season 1"})
    assert res.status_code == 200
    assert res.json()["media_name"] == "Breaking Bad Season 1"


# ===== FEATURE 68: EXCEL EXPORT =====

def test_export_media_excel():
    client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Export Test"})
    res = client.get("/api/v1/excel/export/media")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]


# ===== IS_AVAILABLE FEATURE =====

def test_create_media_default_not_available():
    res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "No Avail Test"})
    assert res.status_code == 201
    data = res.json()
    assert data["is_available"] == "false"
    assert data["available_on"] is None


def test_create_media_with_availability():
    res = client.post("/api/v1/media/", json={
        "media_category": "movies", "media_name": "Avail Test",
        "is_available": "true", "available_on": "Netflix, Amazon Prime"
    })
    assert res.status_code == 201
    data = res.json()
    assert data["is_available"] == "true"
    assert data["available_on"] == "Netflix, Amazon Prime"


def test_update_media_availability():
    create_res = client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Update Avail"})
    mid = create_res.json()["id"]
    res = client.put(f"/api/v1/media/{mid}", json={"is_available": "true", "available_on": "Disney+"})
    assert res.status_code == 200
    assert res.json()["is_available"] == "true"
    assert res.json()["available_on"] == "Disney+"


def test_ui_has_availability_form():
    res = client.get("/")
    assert 'id="fIsAvailable"' in res.text
    assert 'id="fAvailableOn"' in res.text
    assert 'id="availableOnGroup"' in res.text


def test_export_links_excel():
    res = client.get("/api/v1/excel/export/links")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]


def test_export_status_excel():
    res = client.get("/api/v1/excel/export/status")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]


def test_export_media_with_filter():
    res = client.get("/api/v1/excel/export/media?media_category=movies")
    assert res.status_code == 200


# ===== FEATURE 70: TEMPLATES =====

def test_download_media_template():
    res = client.get("/api/v1/excel/templates/media")
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers["content-type"]
    assert "media_import_template" in res.headers["content-disposition"]


def test_download_links_template():
    res = client.get("/api/v1/excel/templates/links")
    assert res.status_code == 200
    assert "links_import_template" in res.headers["content-disposition"]


def test_download_status_template():
    res = client.get("/api/v1/excel/templates/status")
    assert res.status_code == 200
    assert "status_import_template" in res.headers["content-disposition"]


# ===== FEATURE 69: EXCEL IMPORT =====

def test_import_media_excel():
    """Import media from a template Excel file."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["media_category", "media_name", "release_date", "genre", "director", "cast_members", "rating", "review"])
    ws.append(["movies", "imported movie test", "2024-01-01", "Action", "Test Dir", "Actor 1", 8.5, "Great!"])
    ws.append(["shows", "imported show test", "", "Comedy", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post("/api/v1/excel/import/media", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    data = res.json()
    assert data["created"] == 2
    # Check title case was applied
    media_res = client.get("/api/v1/media/?search=Imported+Movie")
    assert media_res.json()["total"] >= 1


def test_import_links_excel():
    """Import links from Excel — needs existing media."""
    client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Link Import Target"})
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["media_name", "platform", "url", "description", "link_status"])
    ws.append(["Link Import Target", "youtube", "https://youtube.com/test", "Test link", "active"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post("/api/v1/excel/import/links", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    assert res.json()["created"] == 1


def test_import_status_excel():
    """Import status from Excel — needs existing media."""
    client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Status Import Target"})
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["media_name", "status", "notes"])
    ws.append(["Status Import Target", "completed", "Finished watching"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post("/api/v1/excel/import/status", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    assert res.json()["created"] == 1


def test_import_invalid_file():
    res = client.post("/api/v1/excel/import/media", files={"file": ("test.txt", b"not excel", "text/plain")})
    assert res.status_code == 400


# ===== UI STRUCTURE TESTS =====

def test_ui_has_dashboard_filters():
    res = client.get("/")
    assert 'id="dashboardCategoryFilter"' in res.text
    assert 'id="dashboardDataFilter"' in res.text


def test_ui_has_export_buttons():
    res = client.get("/")
    assert "exportExcel" in res.text
    assert "downloadTemplate" in res.text
    assert "importExcel" in res.text


def test_ui_version_2_4():
    res = client.get("/")
    assert "v4.0.0" in res.text


# ===== IMPORT PERFORMANCE & DEDUPLICATION TESTS =====

def test_import_media_deduplicates_within_file():
    """Same name appearing multiple times in one file should only be ingested once."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["media_category", "media_name", "release_date", "genre", "director", "cast_members", "rating", "review"])
    ws.append(["webseries", "Duplicate Show", "", "", "", "", "", ""])
    ws.append(["webseries", "Duplicate Show", "", "", "", "", "", ""])  # same name
    ws.append(["webseries", "Duplicate Show", "", "", "", "", "", ""])  # same name again
    ws.append(["webseries", "Unique Show", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post("/api/v1/excel/import/media", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    data = res.json()
    assert data["created"] == 2  # Only 2 unique names
    assert data["summary"]["failed"] == 2  # 2 intra-file duplicates
    assert data["summary"]["total_processed"] == 4
    # Check errors have proper structure
    dup_errors = [e for e in data["errors"] if e["issue_type"] == "duplicate_in_file"]
    assert len(dup_errors) == 2


def test_import_media_skips_existing_db_records():
    """Records already in DB should be reported as duplicates, not crash."""
    client.post("/api/v1/media/", json={"media_category": "movies", "media_name": "Already Exists"})
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["media_category", "media_name", "release_date", "genre", "director", "cast_members", "rating", "review"])
    ws.append(["movies", "Already Exists", "", "", "", "", "", ""])
    ws.append(["movies", "Brand New Movie", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post("/api/v1/excel/import/media", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    data = res.json()
    assert data["created"] == 1
    assert data["summary"]["failed"] == 1
    db_dup_errors = [e for e in data["errors"] if e["issue_type"] == "duplicate"]
    assert len(db_dup_errors) == 1


def test_import_media_returns_summary_structure():
    """Import response should have proper summary structure."""
    from openpyxl import Workbook
    import io
    wb = Workbook()
    ws = wb.active
    ws.append(["media_category", "media_name"])
    ws.append(["movies", "Summary Test Movie"])
    ws.append(["", ""])  # empty row
    ws.append(["", "Missing Category"])  # missing category
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = client.post("/api/v1/excel/import/media", files={"file": ("test.xlsx", buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert res.status_code == 200
    data = res.json()
    assert "summary" in data
    s = data["summary"]
    assert "total_rows" in s
    assert "total_processed" in s
    assert "successful" in s
    assert "failed" in s
    assert "skipped_empty" in s
    assert s["total_rows"] == 3
    assert s["skipped_empty"] >= 1


